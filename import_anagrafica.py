"""Import the Anagrafica (employee master) Excel into the v2 schema.

- Units ("Descrizione unita")  -> Sites (houses); matches existing by accent-
  insensitive name, else creates.
- Role ("Descrizione mansione") -> job_title + JobTitleRecord (Educatore prof.
  variants merged to 'educatore' to reuse the existing house rotations).
- Department derived from the role group (client choice, 2026-07-04).
- Contract: part-time if Perc. part time < 100; monthly hours = weekly x 4.345
  (matches the 104.28 / 130.35 parameter figures).
- Removes the existing sample/test employees first (client choice).

Idempotent-ish: re-running wipes employees + roster and re-imports.
"""

import re
import unicodedata

import openpyxl

from app.core.database import SessionLocal
from app.models.coverage import EmployeeCoverage
from app.models.department import Department
from app.models.employee import ContractType, Employee
from app.models.job_title import JobTitleRecord
from app.models.roster import RosterAssignment
from app.models.site import Site

XLSX = r"C:\Users\SAMAD ALI QURESHI\Downloads\Anagrafica MH al per amir 04.07.2026.xlsx"
HOURS_FACTOR = 4.345  # weekly -> monthly (52.14 / 12)

DEPARTMENTS = {  # name -> code
    "Administrative": "ADM",
    "OSS": "OSS",
    "Auxiliaries": "AUX",
    "COC": "COC",
    "Educatori": "EDC",
    "Notturni": "NTN",
}


def strip_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def norm(s: str) -> str:
    return strip_accents((s or "").strip()).lower()


def dept_for(role_up: str) -> str:
    if "EDUCATORE" in role_up:
        return "Educatori"
    if any(k in role_up for k in ("AUSILIAR", "PULIZIE", "MANUTENT")):
        return "Auxiliaries"
    if any(k in role_up for k in ("NOTTURN", "SORVEGLIAN", "VIGILANTE")):
        return "Notturni"
    if any(k in role_up for k in ("SOCIO ASSIST", "O.S.A", "OSA", "OSS")):
        return "OSS"
    return "Administrative"


def role_key(role: str):
    up = role.strip().upper()
    if up.startswith("EDUCATORE"):
        return "educatore", "Educatore"
    key = re.sub(r"[^a-z0-9]+", "_", role.strip().lower()).strip("_")
    return key, role.strip().title()


def is_junk_unit(u: str) -> bool:
    n = norm(u)
    return (not n) or len(n) < 4 or n in ("ma", "solo mattina")


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Foglio1"]
    rows = [
        [ws.cell(r, c).value for c in range(1, 17)]
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 1).value
    ]
    rows = [r for r in rows if not r[3]]  # active only (no termination date)

    db = SessionLocal()
    try:
        # --- Wipe sample staff + their roster cells ---
        db.query(RosterAssignment).delete(synchronize_session=False)
        db.query(EmployeeCoverage).delete(synchronize_session=False)
        n_removed = db.query(Employee).delete(synchronize_session=False)
        db.flush()

        # --- Departments ---
        dept_by_name = {}
        for name, code in DEPARTMENTS.items():
            d = db.query(Department).filter_by(name=name).first()
            if not d:
                d = Department(name=name, code=code)
                db.add(d)
                db.flush()
            dept_by_name[name] = d

        # --- Sites (match existing by accent-insensitive name) ---
        existing_sites = {norm(s.name): s for s in db.query(Site).all()}
        used_codes = {s.code for s in existing_sites.values()}

        def get_site(unit: str):
            if is_junk_unit(unit):
                return None
            key = norm(unit)
            if key in existing_sites:
                return existing_sites[key]
            name = unit.strip().title()
            base = re.sub(r"[^A-Z0-9]+", "-", strip_accents(name).upper()).strip("-")[:18]
            code = base or "SITE"
            i = 1
            while code in used_codes:
                code = f"{base[:16]}-{i}"
                i += 1
            used_codes.add(code)
            s = Site(name=name, code=code)
            db.add(s)
            db.flush()
            existing_sites[key] = s
            return s

        # --- Job titles ---
        jt_by_key = {j.name: j for j in db.query(JobTitleRecord).all()}

        def get_job_title(role: str) -> str:
            key, label = role_key(role)
            if key not in jt_by_key:
                jt = JobTitleRecord(name=key, label=label)
                db.add(jt)
                db.flush()
                jt_by_key[key] = jt
            return key

        # --- Employees ---
        seen_cf = set()
        inserted = 0
        no_site = []
        for r in rows:
            cognome, nome = (r[0] or "").strip(), (r[1] or "").strip()
            contract = r[5]
            role = (r[8] or "").strip() or "Non specificato"
            perc = r[9]
            weekly = r[10]
            unit = r[12] or ""
            cf = (str(r[13]).strip() or None) if r[13] else None
            if cf and cf in seen_cf:
                cf = None  # avoid unique clash (shouldn't happen)
            if cf:
                seen_cf.add(cf)

            role_up = role.upper()
            dept = dept_by_name[dept_for(role_up)]
            site = get_site(unit)
            if site is None:
                no_site.append(f"{cognome} {nome} (unit={unit!r})")
            jt_key = get_job_title(role)

            try:
                pct = float(perc) if perc is not None else 100.0
            except (TypeError, ValueError):
                pct = 100.0
            ct = ContractType.FULL_TIME if pct >= 100 else ContractType.PART_TIME
            try:
                wk = float(weekly) if weekly else 0.0
            except (TypeError, ValueError):
                wk = 0.0
            monthly = round(wk * HOURS_FACTOR, 2) if wk else 0.0

            db.add(
                Employee(
                    first_name=nome,
                    last_name=cognome,
                    codice_fiscale=cf,
                    department_id=dept.id,
                    site_id=site.id if site else None,
                    job_title=jt_key,
                    location=unit.strip().title() or None,
                    contract_type=ct,
                    monthly_hour_limit=monthly,
                    is_active=True,
                )
            )
            inserted += 1

        db.commit()

        print(f"Removed sample employees: {n_removed}")
        print(f"Imported employees:       {inserted}")
        print(f"Sites total:              {db.query(Site).count()}")
        print(f"Job titles total:         {db.query(JobTitleRecord).count()}")
        if no_site:
            print(f"\nNo site assigned ({len(no_site)}):")
            for x in no_site:
                print(f"  - {x}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
