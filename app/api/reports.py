"""Reporting & export endpoints (v2).

Replaces the v1 attendance/discrepancy report. v2.0 employees receive the
schedule as exported files (req v2.0 §1.7, §4 "Export First"):
  * GET /reports/roster.xlsx  — monthly roster grid as Excel
  * GET /reports/roster.pdf   — monthly roster grid as PDF
  * GET /reports/overtime     — JSON hours/overtime summary (for screen)

No email/SMS, no payroll integration (req v2.0 §1.7).
"""

import calendar
import io
from datetime import date

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.api.deps import require_edit
from app.core.database import get_db
from app.models.employee import Employee
from app.models.roster import RosterAssignment
from app.models.shift_type import ShiftType
from app.models.user import User
from app.services.holidays import holiday_name
from app.services.hours_service import employee_hours_summary, month_bounds

router = APIRouter(prefix="/reports", tags=["Reports & Export"])

# Weekday initials indexed by date.weekday() (Monday = 0 … Sunday = 6).
_WEEKDAY = ["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]


def _build_grid(
    db: Session,
    year: int,
    month: int,
    department_id: int | None,
    job_title: str | None = None,
    site_id: int | None = None,
):
    """Return (employees, days, cell_map) where cell_map[(emp_id, day)] = label."""
    start, end = month_bounds(year, month)
    days = list(range(1, calendar.monthrange(year, month)[1] + 1))

    eq = db.query(Employee).filter(Employee.is_active.is_(True))
    if department_id is not None:
        eq = eq.filter(Employee.department_id == department_id)
    if job_title is not None:
        eq = eq.filter(Employee.job_title == job_title)
    if site_id is not None:
        # Include operators on loan INTO this house (part 2), same as the API.
        onloan_ids = db.query(RosterAssignment.employee_id).filter(
            RosterAssignment.site_id == site_id,
            RosterAssignment.work_date >= start,
            RosterAssignment.work_date <= end,
        )
        eq = eq.filter(
            or_(Employee.site_id == site_id, Employee.id.in_(onloan_ids))
        )
    employees = eq.order_by(Employee.last_name, Employee.first_name).all()

    shift_codes = {s.id: s.code for s in db.query(ShiftType).all()}
    home_site = {e.id: e.site_id for e in employees}

    cells = (
        db.query(RosterAssignment)
        .filter(
            RosterAssignment.work_date >= start,
            RosterAssignment.work_date <= end,
        )
        .all()
    )
    cell_map: dict[tuple[int, int], str] = {}
    for c in cells:
        if site_id is not None:
            emp_home = home_site.get(c.employee_id)
            transferred_out = (
                emp_home == site_id
                and c.site_id is not None
                and c.site_id != site_id
            )
            effective = c.site_id if c.site_id is not None else emp_home
            if transferred_out:
                label = "B2"  # on loan to another house that day
            elif effective == site_id:
                if c.shift_type_id:
                    label = shift_codes.get(c.shift_type_id, "?")
                    if emp_home != site_id:
                        label += "*"  # on loan into this house
                elif c.absence_code:
                    label = c.absence_code.value
                else:
                    label = ""
            else:
                continue  # cell belongs to a house we're not exporting
        else:
            if c.shift_type_id:
                label = shift_codes.get(c.shift_type_id, "?")
            elif c.absence_code:
                label = c.absence_code.value
            else:
                label = ""
        cell_map[(c.employee_id, c.work_date.day)] = label
    return employees, days, cell_map


@router.get("/overtime")
def overtime_report(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    department_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _u: User = Depends(require_edit),
):
    q = db.query(Employee).filter(Employee.is_active.is_(True))
    if department_id is not None:
        q = q.filter(Employee.department_id == department_id)
    return [
        employee_hours_summary(db, e, year, month)
        for e in q.order_by(Employee.last_name).all()
    ]


@router.get("/roster.xlsx")
def export_roster_xlsx(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    department_id: int | None = Query(None),
    job_title: str | None = Query(None),
    site_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _u: User = Depends(require_edit),
):
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill

    employees, days, cell_map = _build_grid(
        db, year, month, department_id, job_title, site_id
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_abbr[month]} {year}"

    header_fill = PatternFill("solid", fgColor="1A7340")
    holiday_fill = PatternFill("solid", fgColor="C0392B")  # Sundays / festività
    header_font = Font(bold=True, color="FFFFFF")
    centre = Alignment(horizontal="center")

    # Two header rows: weekday initial (row 1) over day number (row 2). The
    # "Employee" label spans both. Sundays and festività columns are red.
    ws.merge_cells("A1:A2")
    lbl = ws.cell(row=1, column=1, value="Employee")
    lbl.font = header_font
    lbl.fill = header_fill
    lbl.alignment = Alignment(horizontal="left", vertical="center")

    for i, d in enumerate(days, start=2):
        dt = date(year, month, d)
        hol = holiday_name(dt)
        fill = holiday_fill if (dt.weekday() == 6 or hol) else header_fill
        wd_cell = ws.cell(row=1, column=i, value=_WEEKDAY[dt.weekday()])
        num_cell = ws.cell(row=2, column=i, value=d)
        for c in (wd_cell, num_cell):
            c.font = header_font
            c.fill = fill
            c.alignment = centre
        if hol:
            num_cell.comment = Comment(hol, "Roster")
        ws.column_dimensions[num_cell.column_letter].width = 4

    for r, emp in enumerate(employees, start=3):
        ws.cell(row=r, column=1, value=f"{emp.last_name} {emp.first_name}")
        for i, d in enumerate(days, start=2):
            ws.cell(
                row=r, column=i, value=cell_map.get((emp.id, d), "")
            ).alignment = centre
    ws.column_dimensions["A"].width = 28

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"roster_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/roster.pdf")
def export_roster_pdf(
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    department_id: int | None = Query(None),
    job_title: str | None = Query(None),
    site_id: int | None = Query(None),
    db: Session = Depends(get_db),
    _u: User = Depends(require_edit),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
    from reportlab.lib.styles import getSampleStyleSheet

    employees, days, cell_map = _build_grid(
        db, year, month, department_id, job_title, site_id
    )
    styles = getSampleStyleSheet()

    # Page geometry — landscape A4 with tight margins.
    margin = 12 * mm
    page_w = landscape(A4)[0]
    usable = page_w - 2 * margin
    emp_w = 44 * mm                 # employee-name column
    min_day_w = 9 * mm             # keep day columns readable
    days_per_page = max(1, int((usable - emp_w) // min_day_w))

    # Split the month into column-chunks so nothing is clipped off the page.
    chunks = [days[i : i + days_per_page] for i in range(0, len(days), days_per_page)]

    title = f"Monthly Roster — {calendar.month_name[month]} {year}"
    story = []
    for ci, chunk in enumerate(chunks):
        if ci > 0:
            story.append(PageBreak())
        sub = (
            f"{title}   (days {chunk[0]}–{chunk[-1]})"
            if len(chunks) > 1
            else title
        )
        story.append(Paragraph(sub, styles["Heading2"]))
        story.append(Spacer(1, 6))

        # Two header rows: weekday initials over day numbers. The "Employee"
        # cell spans both (via SPAN below).
        weekday_row = ["Employee"] + [
            _WEEKDAY[date(year, month, d).weekday()] for d in chunk
        ]
        day_row = [""] + [str(d) for d in chunk]
        rows = [weekday_row, day_row]
        for emp in employees:
            rows.append(
                [f"{emp.last_name} {emp.first_name}"]
                + [cell_map.get((emp.id, d), "") for d in chunk]
            )
        day_w = (usable - emp_w) / len(chunk)
        col_widths = [emp_w] + [day_w] * len(chunk)

        # Base style — two green header rows, striped body from row 2.
        style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#1A7340")),
                ("TEXTCOLOR", (0, 0), (-1, 1), colors.white),
                ("SPAN", (0, 0), (0, 1)),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 2), (-1, -1), [colors.white, colors.HexColor("#F3F7F4")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ]
        )
        # Sundays / festività columns → red header.
        for idx, d in enumerate(chunk, start=1):
            dt = date(year, month, d)
            if dt.weekday() == 6 or holiday_name(dt):
                style.add(
                    "BACKGROUND", (idx, 0), (idx, 1), colors.HexColor("#C0392B")
                )
        table = Table(rows, colWidths=col_widths, repeatRows=2)
        table.setStyle(style)
        story.append(table)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
        title=f"Roster {year}-{month:02d}",
    )
    doc.build(story)
    buf.seek(0)
    fname = f"roster_{year}_{month:02d}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
